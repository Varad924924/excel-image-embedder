import streamlit as st
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import zipfile
import io
import os
import tempfile
import pandas as pd

# --- Core Processing Logic (No changes here) ---
def embed_images_to_excel(excel_bytes, image_zip_bytes):
    """
    Embeds images from a zip file into an Excel workbook using specific layout settings.
    """
    with tempfile.TemporaryDirectory() as temp_dir:
        try:
            with zipfile.ZipFile(io.BytesIO(image_zip_bytes), 'r') as zip_ref:
                zip_ref.extractall(temp_dir)

            image_folder = temp_dir
            extracted_items = os.listdir(temp_dir)
            if len(extracted_items) == 1 and os.path.isdir(os.path.join(temp_dir, extracted_items[0])):
                image_folder = os.path.join(temp_dir, extracted_items[0])

        except zipfile.BadZipFile:
            return None, "Error: The uploaded file is not a valid ZIP archive."

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        try:
            header = [cell.value for cell in ws[1]]
            id_col_idx = header.index("Property ID") + 1
            ss_col_idx = header.index("Screenshot") + 1
        except (ValueError, IndexError):
            return None, "Error: 'Property ID' or 'Screenshot' column not found in the Excel file's first row."

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions[get_column_letter(ss_col_idx)].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 50

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 120
            prop_id = ws.cell(row=row, column=id_col_idx).value
            if not prop_id: continue

            image_name = f"part_{prop_id}.jpg"
            image_path = os.path.join(image_folder, image_name)

            if os.path.exists(image_path):
                try:
                    img = XLImage(image_path)
                    img.width = 280
                    img.height = 160
                    cell_anchor = f"{get_column_letter(ss_col_idx)}{row}"
                    ws.add_image(img, cell_anchor)
                except Exception as e:
                    st.warning(f"Could not process image for Property ID {prop_id}: {e}")

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        return output_buffer, None


# --- ⭐ FINAL FIX: The Full Reset Callback Function ⭐ ---
def reset_app():
    """
    This function is called when the reset button is clicked.
    It clears the entire session state to fully restart the app.
    """
    # Get a list of all keys in the session state
    keys_to_delete = list(st.session_state.keys())
    # Delete each key
    for key in keys_to_delete:
        del st.session_state[key]


# --- Streamlit App Main Function ---
def main():
    """Main function to run the Streamlit app."""
    st.set_page_config(layout="wide", page_title="Excel Image Embedder")

    # --- Sidebar Layout ---
    with st.sidebar:
        st.title("🖼️ Excel Image Embedder")
        st.header("Step 1: Upload Your Files")

        uploaded_excel = st.file_uploader(
            "Upload the Excel Workbook (.xlsx)",
            type=["xlsx"],
            key="excel_uploader"
        )
        uploaded_zip = st.file_uploader(
            "Upload the Screenshots ZIP file (.zip)",
            type=["zip"],
            key="zip_uploader"
        )

        # The on_click callback now calls the full reset function
        st.button("🔄 Reset App", on_click=reset_app)

    # --- Main Panel Display ---
    st.header("Instructions")
    st.info(
        "1.  **Upload Files**: Use the sidebar to upload your Excel file and your compressed ZIP file of screenshots.\n"
        "2.  **Preview**: A preview of your Excel data will appear below to confirm it's the correct file.\n"
        "3.  **Embed**: Click the 'Embed Images' button to start the process.\n"
        "4.  **Download**: Once finished, a download button will appear."
    )

    # --- Data Preview Logic ---
    if uploaded_excel:
        st.markdown("---")
        st.subheader("Excel Data Preview")
        try:
            df = pd.read_excel(uploaded_excel)
            st.dataframe(df.head())
        except Exception as e:
            st.error(f"An error occurred while trying to preview the Excel file: {e}")

    st.markdown("---")

    # --- Main Processing and Download Logic ---
    if st.button("🚀 Embed Images into Excel", type="primary"):
        if uploaded_excel is not None and uploaded_zip is not None:
            with st.spinner("Processing... Inserting images into Excel file. ⏳"):
                excel_bytes = uploaded_excel.getvalue()
                zip_bytes = uploaded_zip.getvalue()

                result_buffer, error_message = embed_images_to_excel(excel_bytes, zip_bytes)

            if error_message:
                st.error(error_message)
            else:
                st.success("✅ Success! Images have been embedded.")

                original_filename = os.path.splitext(uploaded_excel.name)[0]
                new_filename = f"{original_filename}_with_images.xlsx"

                st.download_button(
                    label="📥 Download Modified Excel File",
                    data=result_buffer,
                    file_name=new_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("⚠️ Please upload both the Excel file and the ZIP file.")


if __name__ == "__main__":
    main()