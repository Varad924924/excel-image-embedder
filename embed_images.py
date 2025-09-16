import os
import sys
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

def embed_images_in_excel(excel_path, screenshots_dir, output_excel_path=None):
    wb = load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=6):
        screenshot_filename = row[5].value
        if not screenshot_filename:
            continue

        image_path = os.path.join(screenshots_dir, screenshot_filename)
        if os.path.exists(image_path):
            try:
                img = XLImage(image_path)
                img.width = 300
                img.height = 200
                cell_coord = f"F{row[0].row}"
                ws.add_image(img, cell_coord)
                ws.row_dimensions[row[0].row].height = 150
            except Exception as e:
                print(f"Failed to add image for row {row[0].row}: {e}")
        else:
            print(f"Image file not found: {image_path}")

    if not output_excel_path:
        output_excel_path = excel_path.replace(".xlsx", "_with_images.xlsx")

    wb.save(output_excel_path)
    print(f"Saved Excel with embedded images: {output_excel_path}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python embed_images.py <excel_path> <screenshots_dir>")
        sys.exit(1)

    excel_path = sys.argv[1]
    screenshots_dir = sys.argv[2]

    embed_images_in_excel(excel_path, screenshots_dir)
